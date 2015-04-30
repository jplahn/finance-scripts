import csv, openpyxl, os, calendar
from openpyxl.cell import get_column_letter
from openpyxl.styles import Font, Style, Alignment

def open_file_and_return_reader(file):
    print 'Opening csv..'
    return csv.reader(open(file))

def fill_dictionary(dictionary, date, description, amount, category):
    dictionary.setdefault(category, {})
    dictionary[category].setdefault(description, [])

    tup = (date, amount)
    dictionary[category][description].append(tup)

    return dictionary

def add_to_header_dict(dictionary, category, col_index, row_index, total):
    dictionary.setdefault(category, [col_index, row_index, total])

def read_csv_to_dict(reader, dictionary):
    print 'Converting .csv to dictionary..'
    # skip first row
    iter_rows = iter(reader)
    next(iter_rows)
    for row in iter_rows:
        fill_dictionary(dictionary, row[1], row[2], row[3], row[4])

    return dictionary

def add_sheet(wb, sheet_name):
    if not sheet_exists(wb, sheet_name):
        print 'Adding new sheet..'
        wb.create_sheet(title = sheet_name)
        print 'Current sheets: {}'.format(wb.get_sheet_names())

    return wb

def sheet_exists(wb, sheet_name):
    if sheet_name in wb.get_sheet_names():
        return True
    else:
        return False

def add_category_headers(wb, sheet, dictionary, index, header_dict):
    for category in dictionary:
        set_headers(wb, sheet, category, index)
        # Store location of column for each category
        add_to_header_dict(header_dict, category, index, 0, 0)
        index += 3

def add_to_workbook(wb, sheet, dictionary, header_dict):
    for category, record in dictionary.iteritems():
        col_index = header_dict[category][0]
        start_letter = get_column_letter(col_index)
        middle_letter = get_column_letter(col_index + 1)
        end_letter = get_column_letter(col_index + 2)
        row_index = 3
        category_sum = 0
        for description, data in record.iteritems():
            sheet[start_letter + str(row_index)] = description
            sheet[middle_letter + str(row_index)] = data[0][1]
            sheet[end_letter + str(row_index)] = data[0][0]
            row_index += 1
            category_sum += float(data[0][1])
        header_dict[category][1] = row_index + 1
        header_dict[category][2] = round(category_sum, 2)



def set_headers(wb, sheet, category, col_index):
    start_letter = get_column_letter(col_index)
    middle_letter = get_column_letter(col_index + 1)
    end_letter = get_column_letter(col_index + 2)

    sheet.merge_cells(start_letter + '1:' + end_letter + '1')
    sheet[start_letter + '1'] = category
    sheet[start_letter + '1'].font = Font(bold = True, size = 14)
    sheet[start_letter + '1'].alignment = Alignment(horizontal='center')

    sheet[start_letter + '2'] = 'Description'
    sheet[middle_letter + '2'] = 'Amount'
    sheet[end_letter + '2'] = 'Date'

def format_sheet(wb, sheet):
    sheet.freeze_panes = 'A3'

def add_totals_to_sheet(wb, sheet, header_dict):
    for category, record in header_dict.iteritems():
        col = get_column_letter(record[0])
        end_col = get_column_letter(record[0] + 2)
        row = str(record[1])
        sheet.merge_cells(col + row + ':' + end_col + row)
        sheet[col + row] = record[2]
        sheet[col + row].font = Font(bold = True, size = 14)
        sheet[col + row].alignment = Alignment(horizontal='center')

def main():
    reader = open_file_and_return_reader('discover_2015.csv')
    # Creating dictionary
    dictionary = {}
    read_csv_to_dict(reader, dictionary)

    header_dict = {}

    # Creating spreadsheet
    wb = openpyxl.Workbook()
    wb = add_sheet(wb, 'New Sheet')
    sheet = wb.get_sheet_by_name('New Sheet')
    add_category_headers(wb, sheet, dictionary, 1, header_dict)
    add_to_workbook(wb, sheet, dictionary, header_dict)
    format_sheet(wb, sheet)
    add_totals_to_sheet(wb, sheet, header_dict)
    wb.save('test.xlsx')

    print header_dict
    # fill_workbook(wb, dictionary)


main()
